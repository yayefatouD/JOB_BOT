from wttj import scrape_offers
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "wttj_jobs_data.xlsx"
COLUMNS    = ["Intitulé du poste", "Entreprise", "Ville", "Description", "URL"]
COL_WIDTHS = [35, 25, 20, 70, 50]

HEADER_FILL   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ROW_FILL_ODD  = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
ROW_FILL_EVEN = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
CELL_FONT     = Font(name="Arial", size=10)
BORDER        = Border(*[Side(style="thin", color="BDD7EE")] * 0,
                       left=Side(style="thin", color="BDD7EE"),
                       right=Side(style="thin", color="BDD7EE"),
                       top=Side(style="thin", color="BDD7EE"),
                       bottom=Side(style="thin", color="BDD7EE"))


def save_to_excel(offers: list, path: str = EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = "Offres emplois"
    ws.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, offer in enumerate(offers, start=2):
        values = [offer.get("title", ""), offer.get("company", ""),
                  offer.get("location", ""), offer.get("description", ""),
                  offer.get("url", "")]
        fill = ROW_FILL_ODD if row_idx % 2 != 0 else ROW_FILL_EVEN
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = CELL_FONT
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 80

    wb.save(path)
    print(f"{len(offers)} offres sauvegardées dans {path}")


if _name_ == "_main_":
    offers = scrape_offers(job_type="data", location="France")
    save_to_excel(offers)
